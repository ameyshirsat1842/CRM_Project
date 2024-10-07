from datetime import timedelta
from io import BytesIO

import pandas as pd
import pytz
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator
from django.db import IntegrityError
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_GET
from django.views.generic import ListView
from openpyxl.workbook import Workbook

from .forms import SignUpForm, AddRecordForm, AddTicketForm, UpdateRecordForm, AddMeetingRecordForm, PotentialLeadForm, \
    UserUpdateForm, ProfileUpdateForm, CustomerUpdateForm, UpdatePotentialLeadForm, CustomerForm
from .models import Record, Notification, Ticket, MeetingRecord, PotentialLead, Comment, Customer, DeletedRecord, \
    LeadComment
from .utils import verify_otp


def home(request):
    if request.user.is_authenticated:
        now = timezone.now()  # Get the current time
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)  # Start of today
        today_end = now.replace(hour=23, minute=59, second=59, microsecond=999999)  # End of today

        # Get data specific to the logged-in user
        total_leads = Record.objects.filter(assigned_to=request.user).count()
        total_clients = Record.objects.filter(assigned_to=request.user).values('client_name').distinct().count()
        open_tickets = Ticket.objects.filter(created_by=request.user, status='Open').count()
        closed_deals = Record.objects.filter(assigned_to=request.user, classification='dead').count()

        # Get recent activities specific to the logged-in user
        recent_leads = Record.objects.filter(assigned_to=request.user).order_by('-created_at')[:5]
        recent_tickets = Ticket.objects.filter(created_by=request.user).order_by('-created_at')[:5]

        # Get upcoming meetings (from tomorrow onwards)
        upcoming_meetings = Record.objects.filter(
            assigned_to=request.user, follow_up_date__gt=today_end
        ).order_by('follow_up_date')[:5]

        # Meetings today
        meetings_today = Record.objects.filter(
            assigned_to=request.user,
            follow_up_date__gte=today_start,
            follow_up_date__lte=today_end
        ).order_by('follow_up_date')[:5]

        # Get overdue leads: leads with follow-up dates in the past (before today)
        overdues = Record.objects.filter(
            assigned_to=request.user,
            follow_up_date__lt=now,
            classification='in_progress',
            follow_up_attended=False,
            is_converted=False
        )

        # Notifications for the logged-in user
        notifications = Notification.objects.unread_for_user(request.user)

        context = {
            'total_leads': total_leads,
            'total_clients': total_clients,
            'open_tickets': open_tickets,
            'closed_deals': closed_deals,
            'recent_leads': recent_leads,
            'recent_tickets': recent_tickets,
            'upcoming_meetings': upcoming_meetings,
            'meetings_today': meetings_today,
            'notifications': notifications,
            'overdues': overdues,
        }

        return render(request, 'home.html', context)
    else:
        return redirect('login')


def login_user(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "You are logged in")
            return redirect('home')
        else:
            messages.error(request, "Invalid User")
            return redirect('login')
    else:
        return render(request, 'login.html')  # Ensure this is 'login.html'


def logout_user(request):
    logout(request)
    messages.success(request, "You have been logged out")
    return redirect('login')  # Redirect to the login page after logout


def register_user(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False  # Deactivate the account until email verification
            user.save()

            messages.success(request, "Registration successful! Please check your email for the OTP.")
            return redirect('otp_verify', user_id=user.id)  # Redirect to OTP verification page
        else:
            errors = form.errors.as_json()
            return JsonResponse({'errors': errors}, status=400)
    else:
        form = SignUpForm()

    return render(request, 'register.html', {'form': form})


def otp_verify(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        otp = request.POST.get('otp')
        if verify_otp(user, otp):
            user.is_active = True
            user.save()
            login(request, user)
            messages.success(request, "Your account has been activated. You are now logged in.")
            return redirect('home')
        else:
            messages.error(request, "Invalid OTP. Please try again.")

    return render(request, 'otp_verify.html', {'user_id': user.id})


def update_user_info(request):
    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = ProfileUpdateForm(request.POST, instance=request.user.profile)

        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, 'Your profile has been updated successfully!')
            return redirect('update_user_info')
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = ProfileUpdateForm(instance=request.user.profile)

    return render(request, 'update_user_info.html', {
        'user_form': user_form,
        'profile_form': profile_form,
    })


def customer_record(request, pk):
    if request.user.is_authenticated:
        customer_rec = get_object_or_404(Record, id=pk)
        meeting_records = MeetingRecord.objects.filter(record=customer_rec)
        can_delete = customer_rec.created_by == request.user
        return render(request, 'record.html', {
            'customer_record': customer_rec,
            'meeting_records': meeting_records,
            'can_delete': can_delete
        })
    else:
        messages.error(request, "You must be logged in to view the record")
        return redirect('home')


def delete_record(request, pk):
    if not request.user.is_authenticated:
        messages.error(request, "You must be logged in to delete records.")
        return redirect('login')

    if not request.user.is_staff:
        messages.error(request, "You do not have permission to delete this record.")
        return redirect('record', pk=pk)

    record = get_object_or_404(Record, pk=pk)

    if request.method == 'POST':
        deletion_reason = request.POST.get('deletion_reason', None)

        if not deletion_reason:
            messages.error(request, "You must provide a reason for deletion.")
            return render(request, 'delete_record.html', {'record': record})

        # Temporarily set the deletion_reason to the record instance
        record.deletion_reason = deletion_reason

        record.delete()
        messages.success(request, "Record deleted successfully.")
        return redirect('leads')

    return render(request, 'delete_record.html', {'record': record})


def deletion_history(request):
    history = DeletedRecord.objects.all()
    return render(request, 'deleted_records.html', {'history': history})


def deleted_record_detail(request, pk):
    record = get_object_or_404(DeletedRecord, pk=pk)
    return render(request, 'deleted_record_detail.html', {'record': record})


def add_record(request):
    if request.method == 'POST':
        form = AddRecordForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            phone = form.cleaned_data.get("phone")
            email = form.cleaned_data.get("email")

            # Check if a record with the same phone number and email exists
            existing_record = Record.objects.filter(phone=phone, email=email).first()
            if existing_record:
                messages.error(request, "A record with this phone number and email already exists.")
                return render(request, 'add_record.html', {'form': form, 'users': User.objects.all()})

            # Create and save the record
            record = form.save(commit=False)
            record.created_by = request.user
            record.last_modified_by = request.user

            # Convert follow_up_date to timezone-aware datetime if it's naive
            follow_up_date = form.cleaned_data['follow_up_date']
            if follow_up_date and timezone.is_naive(follow_up_date):
                follow_up_date = timezone.make_aware(follow_up_date, timezone.get_current_timezone())
                record.follow_up_date = follow_up_date

            # Handle attachments if present
            if 'attachments' in request.FILES:
                record.attachments = request.FILES['attachments']

            record.save()

            messages.success(request, "Record added successfully!")
            return redirect('leads')
        else:
            messages.error(request, "There was an error with the form.")
    else:
        form = AddRecordForm(user=request.user)

    users = User.objects.all()
    return render(request, 'add_record.html', {'form': form, 'users': users})


def update_record(request, pk):
    record = get_object_or_404(Record, pk=pk)

    if request.method == 'POST':
        form = UpdateRecordForm(request.POST, request.FILES, instance=record)

        if form.is_valid():
            updated_record = form.save(commit=False)
            updated_record.last_modified_by = request.user

            # Convert follow_up_date to timezone-aware datetime if it's naive
            follow_up_date = form.cleaned_data['follow_up_date']
            if follow_up_date and timezone.is_naive(follow_up_date):
                follow_up_date = timezone.make_aware(follow_up_date, timezone.get_current_timezone())
                updated_record.follow_up_date = follow_up_date

            # Handle attachments if present, otherwise keep the existing file
            if 'attachments' in request.FILES:
                updated_record.attachments = request.FILES['attachments']
            else:
                # Preserve the existing file if no new file is uploaded
                updated_record.attachments = record.attachments

            updated_record.save()

            # Process additional comments from the dynamically added comment fields
            additional_comments = request.POST.getlist('additional_comments[]')
            for comment in additional_comments:
                if comment.strip():  # Only save non-empty comments
                    Comment.objects.create(record=record, text=comment, user=request.user)

            messages.success(request, "Record updated successfully!")
            return redirect('leads')
        else:
            messages.error(request, "Error updating the record. Please correct the errors below.")
    else:
        form = UpdateRecordForm(instance=record)

    users = User.objects.all()
    return render(request, 'update_record.html', {'form': form, 'record': record, 'users': users})


def send_notification_to_user(user, message, link_url=None):
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        f'notifications_{user.id}',
        {
            'type': 'send_notification',
            'notification': message,
            'link_url': link_url  # Optional URL if available

        }
    )


def notifications_view(request):
    if request.user.is_authenticated:
        notifications = Notification.objects.unread_for_user(request.user)
        notifications_count = notifications.count()
    else:
        notifications = []
        notifications_count = 0

    return render(request, 'notifications.html', {
        'notifications': notifications,
        'notifications_count': notifications_count
    })


def notification_detail(request, pk):
    notification = get_object_or_404(Notification, pk=pk)
    notification.is_read = True
    notification.save()
    return render(request, 'notification_detail.html', {'notification': notification})


@require_GET
def mark_notification_as_read(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.is_read = True
    notification.save()
    return JsonResponse({'status': 'success'})


def tickets(request):
    # Get search and filter parameters from the request
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')

    # Query tickets with search and status filtering
    tickets = Ticket.objects.select_related('customer').all()

    if search_query:
        tickets = tickets.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(company_name__icontains=search_query) |
            Q(contact_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    if status_filter in ['Inactive', 'In Progress', 'Closed']:
        tickets = tickets.filter(status=status_filter)

    return render(request, 'tickets.html', {'tickets': tickets, 'search_query': search_query, 'status_filter': status_filter})


def add_ticket(request):
    if request.method == "POST":
        form = AddTicketForm(request.POST, request.FILES)  # Include request.FILES if there's a file upload
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by = request.user  # Set the creator as the logged-in user

            if ticket.customer:
                customer = ticket.customer
                ticket.company_name = customer.company
                ticket.contact_name = customer.client_name
                ticket.phone = customer.phone
                ticket.email = customer.email

            ticket.save()
            messages.success(request, "Ticket added successfully!")
            return redirect('tickets')
        else:
            # Log or print form errors for debugging
            print(form.errors)
            messages.error(request, "Error adding ticket. Please check your inputs.")
    else:
        form = AddTicketForm()
    customers = Customer.objects.all()  # Pass customers to the template
    return render(request, 'add_ticket.html', {'form': form, 'customers': customers})


def update_ticket(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.method == "POST":
        form = AddTicketForm(request.POST, request.FILES, instance=ticket)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.last_modified_by = request.user  # Set the last_modified_by to the current user
            ticket.save()
            messages.success(request, "Ticket updated successfully!")
            return redirect('tickets')
        else:
            messages.error(request, "Error updating ticket. Please check your inputs.")
    else:
        form = AddTicketForm(instance=ticket)
    return render(request, 'update_ticket.html', {'form': form})


def ticket_detail(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    return render(request, 'ticket_detail.html', {'ticket': ticket})


def add_meeting_record(request, pk):
    record = get_object_or_404(Record, pk=pk)
    if request.method == 'POST':
        form = AddMeetingRecordForm(request.POST)
        if form.is_valid():
            meeting_record = form.save(commit=False)
            meeting_record.record = record
            meeting_record.created_by = request.user

            # Convert follow_up_date to timezone-aware datetime if it's naive
            follow_up_date = form.cleaned_data.get('follow_up_date')
            if follow_up_date and timezone.is_naive(follow_up_date):
                follow_up_date = timezone.make_aware(follow_up_date, timezone.get_current_timezone())
                meeting_record.follow_up_date = follow_up_date

            meeting_record.save()
            form.save_m2m()  # This saves the ManyToManyField relationships, including attendees
            messages.success(request, "Meeting record added successfully!")
            return redirect('record', pk=record.pk)
    else:
        form = AddMeetingRecordForm()

    return render(request, 'add_meeting_record.html', {'form': form, 'record': record})


def update_meeting_record(request, pk):
    record = get_object_or_404(MeetingRecord, pk=pk)
    if request.method == 'POST':
        form = AddMeetingRecordForm(request.POST, instance=record)
        if form.is_valid():
            meeting_record = form.save(commit=False)

            # Convert follow_up_date to timezone-aware datetime if it's naive
            follow_up_date = form.cleaned_data.get('follow_up_date')
            if follow_up_date and timezone.is_naive(follow_up_date):
                follow_up_date = timezone.make_aware(follow_up_date, timezone.get_current_timezone())
                meeting_record.follow_up_date = follow_up_date

            meeting_record.save()
            form.save_m2m()  # This saves the ManyToManyField relationships, including attendees
            messages.success(request, "Meeting record updated successfully!")
            return redirect('meeting_records')
    else:
        form = AddMeetingRecordForm(instance=record)

    return render(request, 'update_meeting_record.html', {'form': form})


def meeting_records(request, pk):
    record = get_object_or_404(Record, pk=pk)
    records = record.meeting_records.all()
    return render(request, 'meeting_records.html', {'records': records, 'record': record})


class MeetingRecordListView(ListView):
    model = MeetingRecord
    template_name = 'meeting_records.html'
    context_object_name = 'records'
    paginate_by = 10  # Adjust as needed

    def get_queryset(self):
        queryset = super().get_queryset().order_by('-created_at')
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(meeting_partner__icontains=query) |
                Q(products_discussed_partner__icontains=query) |
                Q(products_discussed_company__icontains=query) |
                Q(conclusion__icontains=query) |
                Q(follow_up_date__icontains=query)
            )
        return queryset


def delete_meeting_record(request, pk):
    record = get_object_or_404(MeetingRecord, pk=pk)
    if request.method == 'POST':
        record.delete()
        messages.success(request, "Meeting record deleted successfully!")
        return redirect('meeting_records')
    return render(request, 'delete_meeting_record.html', {'record': record})


def add_potential_lead(request):
    if request.method == 'POST':
        form = PotentialLeadForm(request.POST)
        if form.is_valid():
            potential_lead = form.save(commit=False)
            potential_lead.created_by = request.user
            potential_lead.save()
            messages.success(request, "Potential lead added successfully!")
            return redirect('potential_leads')
        else:
            messages.error(request, "Error adding potential lead. Please check the form.")
    else:
        form = PotentialLeadForm()
    return render(request, 'add_potential_lead.html', {'form': form})


def update_potential_lead(request, lead_id):
    lead = get_object_or_404(PotentialLead, id=lead_id)

    if request.method == "POST":
        form = UpdatePotentialLeadForm(request.POST, instance=lead)
        if form.is_valid():
            # Save the main lead form
            form.save()

            # Handle additional comments if any
            additional_comments = request.POST.getlist('additional_comments[]')
            for comment_text in additional_comments:
                if comment_text.strip():  # Make sure it's not empty
                    # Save each additional comment as a new LeadComment object
                    LeadComment.objects.create(lead=lead, user=request.user, comment=comment_text)

            messages.success(request, "Connection updated successfully!")
            return redirect('potential_leads')  # Redirect after updating
        else:
            messages.error(request, "Error updating connection. Please check the form.")
    else:
        form = UpdatePotentialLeadForm(instance=lead)

    # Fetch existing comments to show in the modal, not in the editable form
    existing_comments = LeadComment.objects.filter(lead=lead)

    return render(request, 'update_potential_lead.html', {'form': form, 'existing_comments': existing_comments})


def potential_leads(request):
    leads = PotentialLead.objects.all()
    return render(request, 'potential_leads.html', {'leads': leads})


def potential_lead_detail(request, potential_lead_id):
    connection = get_object_or_404(PotentialLead, id=potential_lead_id)
    additional_comments = LeadComment.objects.filter(lead=connection).order_by('-timestamp')
    return render(request, 'connection_detail.html', {
        'connection': connection,
        'additional_comments': additional_comments
    })


def leads_view(request):
    if request.user.is_authenticated:
        search_query = request.GET.get('search', '')
        classification = request.GET.get('classification', '')
        priority = request.GET.get('priority', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')

        records = Record.objects.filter(is_converted=False).order_by('-created_at')

        # Apply search filter
        if search_query:
            records = records.filter(
                Q(company__icontains=search_query) |
                Q(client_name__icontains=search_query) |
                Q(dept_name__icontains=search_query) |
                Q(phone__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(city__icontains=search_query) |
                Q(assigned_to__username__icontains=search_query) |
                Q(comments__icontains=search_query) |
                Q(remarks__icontains=search_query)
            )

        # Apply classification filter
        if classification and classification != 'all':
            records = records.filter(classification=classification)

        # Apply priority filter
        if priority:
            records = records.filter(priority=priority)

        # Apply date filter
        if start_date and end_date:
            records = records.filter(created_at__range=[start_date, end_date])

        paginator = Paginator(records, 10)  # Show 10 leads per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'page_obj': page_obj,
            'search_query': search_query,
            'classification': classification,
            'priority': priority,
            'start_date': start_date,
            'end_date': end_date,
        }
        return render(request, 'leads.html', context)
    else:
        return redirect('login')


def move_to_main_leads(request, lead_id):
    lead = PotentialLead.objects.get(id=lead_id)
    main_lead = Record(
        company=lead.company,
        client_name=lead.client_name,
        phone=lead.phone,
        email=lead.email,
        comments=lead.initial_comments,
        created_by=lead.created_by
    )
    main_lead.save()
    lead.delete()
    messages.success(request, "Lead moved to main leads successfully!")
    return redirect('leads')


def customers(request):
    search_query = request.GET.get('search', '')
    classification = request.GET.get('classification', '')
    filter_option = request.GET.get('filter', '')

    # Base queryset for filtering customers created by the logged-in user
    customers_list = Customer.objects.all()

    # Apply search filter if provided
    if search_query:
        customers_list = customers_list.filter(client_name__icontains=search_query)

    # Apply status filter if provided
    if classification:
        customers_list = customers_list.filter(classification=classification)

    # Filter customers assigned to the current user if the filter_option is set to 'assigned_to_me'
    if filter_option == 'assigned_to_me':
        customers_list = customers_list.filter(assigned_to=request.user)

    # Add ordering to the queryset
    customers_list = customers_list.order_by('-created_at')

    # Paginate the results
    paginator = Paginator(customers_list, 10)  # Show 10 customers per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'customers.html', {'page_obj': page_obj})


def add_customer(request):
    if request.method == "POST":
        form = CustomerForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('customers')  # Redirect to a list of customers or any relevant page
    else:
        form = CustomerForm()

    return render(request, 'add_customer.html', {'form': form})


def move_to_customers(request, record_id):
    try:
        # Retrieve the record
        record = get_object_or_404(Record, id=record_id)

        # Check if a Customer with the same email already exists to avoid duplicates
        if Customer.objects.filter(email=record.email).exists():
            messages.error(request, f"A customer with email {record.email} already exists.")
            return redirect('customers')

        # Map Record classification to Customer classification
        if record.classification == 'assigned':
            customer_classification = 'active'
        elif record.classification == 'in_progress':
            customer_classification = 'active'
        elif record.classification == 'dead':
            customer_classification = 'inactive'
        else:
            customer_classification = 'Converted'  # Default or fallback classification

        # Create a new Customer instance from the Record
        customer = Customer(
            company=record.company,
            client_name=record.client_name,
            phone=record.phone,
            email=record.email,
            address=record.address,
            city=record.city,
            dept_name=record.dept_name,
            assigned_to=record.assigned_to,
            lead_source=record.lead_source,
            remarks=record.remarks,
            comments=record.comments,
            created_by=record.created_by,
            created_at=record.created_at,
            last_modified_by=record.last_modified_by,
            classification=customer_classification,  # Use the mapped classification
            bank_details='',  # Default value or blank if not yet provided
            gst_number=''  # Default value or blank if not yet provided
        )

        # Save the new Customer to the database
        customer.save()

        # Optionally delete the record or perform other actions
        record.is_converted = True
        record.save()

        messages.success(request, "Record moved to customers successfully!")
        return redirect('customers')

    except IntegrityError:
        messages.error(request, "An integrity error occurred. There was an issue with moving the record to customers.")
        return redirect('leads')

    except Record.DoesNotExist:
        messages.error(request, "The lead you are trying to move does not exist.")
        return redirect('leads')

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        return redirect('leads')


def customer_detail(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    return render(request, 'customer_detail.html', {'customer': customer})


def update_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)

    if request.method == 'POST':
        form = CustomerUpdateForm(request.POST, request.FILES, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, "Customer updated successfully!")
            return redirect('customers')
    else:
        form = CustomerUpdateForm(instance=customer)

    return render(request, 'update_customer.html', {'form': form, 'customer': customer})


def delete_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer.delete()
        messages.success(request, "Customer deleted successfully!")
        return redirect('customers')

    return render(request, 'delete_customer.html', {'customer': customer})


def assign_lead(request, pk):
    lead = get_object_or_404(Record, pk=pk)
    if request.method == 'POST':
        lead.assigned_to = request.user
        lead.save()

        message = f"Lead assigned to you: {lead.company}"
        send_notification_to_user(request.user, message)

        messages.success(request, "Lead assigned successfully!")
        return redirect('leads')


def export_record_to_excel(request, record_id):
    try:
        # Fetch the record by ID
        record = Record.objects.get(pk=record_id)

        # Convert timezone-aware datetime to timezone-naive
        if record.created_at.tzinfo is not None:
            created_at = record.created_at.astimezone(pytz.utc).replace(tzinfo=None)
        else:
            created_at = record.created_at

        follow_up_date = record.follow_up_date if record.follow_up_date else None

        # Convert the record to a DataFrame
        data = {
            'ID': [record.id],
            'Company': [record.company],
            'Client': [record.client_name],
            'Department': [record.dept_name],
            'Phone': [record.phone],
            'Email': [record.email],
            'City': [record.city],
            'Address': [record.address],
            'Follow-Up': [follow_up_date],
            'Comments': [record.comments],
            'Meeting Type': [record.remarks],
            'Attachments': [record.attachments.name if record.attachments else 'No attachments'],
            'Assigned To': [record.assigned_to.username if record.assigned_to else 'N/A'],
            'Created By': [record.created_by.username if record.created_by else 'N/A'],
            'Event Details': [record.social_media_details],
            'Status': [record.classification],
            'Lead Source': [record.lead_source],
            'Created At': [created_at],
        }
        df = pd.DataFrame(data)

        # Create a BytesIO buffer
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

        buffer.seek(0)

        # Create the HttpResponse object
        response = HttpResponse(buffer,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=record_{record_id}.xlsx'

        return response
    except Record.DoesNotExist:
        return HttpResponse("Record not found.", status=404)


def import_records_from_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        fs = FileSystemStorage()
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)

        try:
            # Read the Excel file into a DataFrame
            df = pd.read_excel(file_path)

            # Clean up column names
            df.columns = df.columns.str.strip()

            # Print columns for debugging
            print("Detected columns:", df.columns.tolist())

            # Update the expected columns list according to the actual columns
            expected_columns = [
                'ID', 'Company', 'Client', 'Department', 'Phone', 'Email',
                'City', 'Address', 'Follow-Up', 'Comments', 'Meeting Type',
                'Event Details', 'Lead Source', 'Assigned To', 'Status',
                'Created', 'Follow-Up'
            ]

            missing_columns = [col for col in expected_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(f'Excel file is missing the following columns: {", ".join(missing_columns)}')

            # Iterate over the rows in the DataFrame and create Record objects
            for index, row in df.iterrows():
                try:
                    created_at = pd.to_datetime(row.get('Created'), errors='coerce')
                    follow_up_date = pd.to_datetime(row.get('Follow-Up'), errors='coerce')

                    # Create or update the Record object
                    record = Record(
                        created_at=created_at,
                        company=row.get('Company'),
                        client_name=row.get('Client'),
                        dept_name=row.get('Department'),
                        phone=row.get('Phone'),
                        email=row.get('Email'),
                        city=row.get('City'),
                        address=row.get('Address'),
                        follow_up_date=follow_up_date,
                        comments=row.get('Comments'),
                        remarks=row.get('Meeting Type'),
                        attachments=None,  # Handle file attachments separately if needed
                        assigned_to=User.objects.get(username=row.get('Assigned To')) if pd.notna(
                            row.get('Assigned To')) else None,
                        social_media_details=row.get('Event Details'),
                        classification=row.get('Status'),
                        lead_source=row.get('Lead Source'),
                    )

                    # Save the record
                    record.save()

                except ValidationError as ve:
                    print(f"Validation error for row {index}: {ve}")
                except Exception as e:
                    print(f"Error saving row {index}: {e}")

            # Redirect to the leads view with a success message
            messages.success(request, 'Records imported successfully.')
            return redirect('leads')

        except Exception as e:
            # Print the exception to the console and display an error message
            print("Error processing file:", str(e))
            messages.error(request, 'Error processing file. Please ensure the file is in the correct format.')
            return redirect('import_leads')

    return render(request, 'import_leads.html')


def export_leads(request):
    search_query = request.GET.get('search', '')
    classification_filter = request.GET.get('classification', '')
    assigned_filter = request.GET.get('filter', '')
    priority_filter = request.GET.get('priority', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    queryset = Record.objects.filter(is_converted=False)

    # Apply search filter
    if search_query:
        queryset = queryset.filter(
            Q(company__icontains=search_query) |
            Q(client_name__icontains=search_query) |
            Q(dept_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(comments__icontains=search_query) |
            Q(remarks__icontains=search_query) |
            Q(social_media_details__icontains=search_query) |
            Q(lead_source__icontains=search_query)
        )

    # Apply classification filter
    if classification_filter:
        queryset = queryset.filter(classification=classification_filter)

    # Apply 'Assigned to Me' filter
    if assigned_filter == 'assigned_to_me':
        queryset = queryset.filter(assigned_to=request.user)

    # Apply priority filter
    if priority_filter:
        queryset = queryset.filter(priority=priority_filter)

    # Apply date filter
    if start_date and end_date:
        queryset = queryset.filter(created_at__range=[start_date, end_date])

    # Create Excel file
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Leads'

    headers = [
        'ID', 'Company', 'Client', 'Department', 'Phone', 'Email',
        'City', 'Address', 'Comments', 'Meeting type', 'Event Details',
        'Lead Source', 'Assigned To', 'Status', 'Priority', 'Created', 'Follow-Up'
    ]
    worksheet.append(headers)

    for record in queryset:
        worksheet.append([
            record.id, record.company, record.client_name, record.dept_name,
            record.phone, record.email, record.city, record.address,
            record.comments, record.remarks, record.social_media_details,
            record.lead_source, record.assigned_to.username if record.assigned_to else 'N/A',
            record.get_classification_display(),
            record.get_priority_display(),
            record.created_at.strftime('%Y-%m-%d'),
            record.follow_up_date.strftime('%Y-%m-%d') if record.follow_up_date else ''
        ])

    # Prepare the response
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = HttpResponse(stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=leads.xlsx'
    return response


def export_customers(request):
    search_query = request.GET.get('search', '')
    classification_filter = request.GET.get('classification', '')
    assigned_filter = request.GET.get('filter', '')

    queryset = Customer.objects.all()

    if search_query:
        queryset = queryset.filter(
            Q(company__icontains=search_query) |
            Q(client_name__icontains=search_query) |
            Q(dept_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(comments__icontains=search_query) |
            Q(remarks__icontains=search_query)
        )

    if classification_filter:
        queryset = queryset.filter(classification=classification_filter)

    if assigned_filter == 'assigned_to_me':
        queryset = queryset.filter(assigned_to=request.user)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Customers'

    headers = [
        'ID', 'Company', 'Client', 'Department', 'Phone', 'Email',
        'City', 'Address', 'Comments', 'Remarks',
        'Lead Source', 'Assigned To', 'Status', 'Created'
    ]
    worksheet.append(headers)

    for customer in queryset:
        worksheet.append([
            customer.id, customer.company, customer.client_name, customer.dept_name,
            customer.phone, customer.email, customer.city, customer.address,
            customer.comments, customer.remarks,
            customer.lead_source, customer.assigned_to.username if customer.assigned_to else 'N/A',
            customer.get_classification_display(),
            customer.created_at.strftime('%Y-%m-%d')
        ])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = HttpResponse(stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customers.xlsx'
    return response


def export_deleted_records(request):
    # Query all deleted records
    deleted_records = DeletedRecord.objects.all()

    # Create a new workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Deleted Records'

    # Define headers, including the deletion reason
    headers = [
        'Record ID', 'Client Name', 'Company', 'Assigned To',
        'Deleted By', 'Deleted At', 'Reason'
    ]
    worksheet.append(headers)

    # Add data rows for each deleted record
    for record in deleted_records:
        worksheet.append([
            record.record_id,
            record.client_name,
            record.company,
            record.assigned_to.username if record.assigned_to else 'Unassigned',
            record.deleted_by.username if record.deleted_by else 'Unknown',
            record.deleted_at.strftime('%Y-%m-%d %H:%M:%S'),
            record.deletion_reason or 'No reason provided'  # Include deletion reason
        ])

    # Save workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create an HTTP response with the Excel file
    response = HttpResponse(stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=deleted_records.xlsx'
    return response


def export_meeting_records(request):
    # Query all meeting records
    meeting_records = MeetingRecord.objects.all()

    # Create a new workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Meeting Records'

    # Define headers
    headers = [
        'Meeting ID', 'Client Name', 'Company', 'Meeting Partner', 'Products Discussed with Partner',
        'Products Discussed with Company', 'Conclusion', 'Follow-Up Date', 'Created By',
        'Created At', 'Speaker', 'Attendees', 'Location'
    ]
    worksheet.append(headers)

    # Add data rows for each meeting record
    for meeting in meeting_records:
        attendees = ', '.join([attendee.username for attendee in meeting.attendees.all()])  # List all attendees
        worksheet.append([
            meeting.id,
            meeting.record.client_name if meeting.record else 'N/A',
            meeting.record.company if meeting.record else 'N/A',
            meeting.meeting_partner,
            meeting.products_discussed_partner,
            meeting.products_discussed_company,
            meeting.conclusion,
            meeting.follow_up_date.strftime('%Y-%m-%d %H:%M:%S') if meeting.follow_up_date else 'N/A',
            meeting.created_by.username if meeting.created_by else 'N/A',
            meeting.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            meeting.speaker.username if meeting.speaker else 'N/A',
            attendees,
            meeting.meeting_location,
        ])

    # Save workbook to a BytesIO stream
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Create an HTTP response with the Excel file
    response = HttpResponse(stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=meeting_records.xlsx'
    return response


def settings_view(request):
    return render(request, 'settings.html')


def master_database(request):
    # Get search query from the request
    search_query = request.GET.get('search', '')

    # Fetch all leads, users, customers, tickets, and deleted records
    leads = Record.objects.filter(is_converted=False)
    users = User.objects.all()
    customers = Customer.objects.all()
    deleted_records = DeletedRecord.objects.all()
    tickets = Ticket.objects.all()  # Fetch all tickets

    # Filter based on the search query if provided
    if search_query:
        leads = leads.filter(
            Q(company__icontains=search_query) |
            Q(client_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query)
        )
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )
        customers = customers.filter(
            Q(company__icontains=search_query) |
            Q(client_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query)
        )
        deleted_records = deleted_records.filter(
            Q(company__icontains=search_query) |
            Q(client_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query)
        )
        tickets = tickets.filter(  # Apply search filter to tickets
            Q(company_name__icontains=search_query) |
            Q(contact_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    context = {
        'leads': leads,
        'users': users,
        'customers': customers,
        'deleted_records': deleted_records,
        'tickets': tickets,  # Include tickets in the context
        'search_query': search_query
    }

    return render(request, 'master_database.html', context)


def delete_account(request):
    if request.method == 'POST':
        user = request.user
        user.delete()
        messages.success(request, "Your account has been deleted successfully.")
        logout(request)  # Log out the user after deletion
        return redirect('home')  # Redirect to home page after deletion
    return redirect('update_info')  # Redirect back if accessed via GET


def reports(request):
    # Get selected filters from the request
    selected_period = request.GET.get('period', 'all')
    selected_classification = request.GET.get('classification', 'all')
    selected_user = request.GET.get('user', 'all')
    custom_start_date = request.GET.get('start_date', '')
    custom_end_date = request.GET.get('end_date', '')

    # Define the time filter based on selected_period
    if selected_period == 'custom':
        # Custom period selected
        start_date = timezone.datetime.strptime(custom_start_date, '%Y-%m-%d').date() if custom_start_date else None
        end_date = timezone.datetime.strptime(custom_end_date, '%Y-%m-%d').date() if custom_end_date else None
    elif selected_period == 'last_3_days':
        start_date = timezone.now() - timedelta(days=3)
        end_date = None
    elif selected_period == 'last_week':
        start_date = timezone.now() - timedelta(weeks=1)
        end_date = None
    elif selected_period == 'last_month':
        start_date = timezone.now() - timedelta(days=30)
        end_date = None
    else:
        start_date = None
        end_date = None  # No date filter for 'all'

    # Query for all records initially
    records = Record.objects.filter(is_converted=False)
    customers = Customer.objects.all()

    # Apply time filter if a start date is defined
    if start_date:
        records = records.filter(created_at__gte=start_date)
        customers = customers.filter(created_at__gte=start_date)
    if end_date:
        records = records.filter(created_at__lte=end_date)
        customers = customers.filter(created_at__lte=end_date)

    # Apply classification filter if a specific classification is selected
    if selected_classification == 'over_due':
        records = records.filter(follow_up_date__lt=timezone.now(), classification='in_progress', follow_up_attended=False)
    elif selected_classification != 'all':
        records = records.filter(classification=selected_classification)
        customers = customers.filter(classification=selected_classification)

    # Apply user filter if a specific user is selected
    if selected_user != 'all':
        records = records.filter(assigned_to_id=selected_user)
        customers = customers.filter(assigned_to_id=selected_user)  # Filter customers by assigned user

    # Calculate counts for different categories
    overdues_count = records.filter(follow_up_date__lt=timezone.now(), classification='in_progress', follow_up_attended=False).count()
    leads_count = records.count()  # Total leads count
    converted_count = customers.filter(classification='active').count()  # Example classification filter for customers

    context = {
        'records': records,
        'customers': customers,
        'users': User.objects.all(),
        'selected_period': selected_period,
        'selected_classification': selected_classification,
        'selected_user': selected_user,
        'custom_start_date': custom_start_date,
        'custom_end_date': custom_end_date,
        'overdues_count': overdues_count,
        'leads_count': leads_count,
        'converted_count': converted_count,
    }
    return render(request, 'reports.html', context)


def admin_dashboard(request):
    # Aggregating data for the dashboard
    total_leads = Record.objects.count()
    total_customers = Customer.objects.count()
    total_users = User.objects.count()

    # For example, calculate leads from the last 30 days
    recent_leads = Record.objects.filter(created_at__gte=timezone.now() - timedelta(days=30)).count()
    overdue_leads = Record.objects.filter(follow_up_date__lt=timezone.now(), classification='in_progress').count()

    # New: User-wise leads
    users = User.objects.all()
    user_leads = []
    for user in users:
        leads = Record.objects.filter(assigned_to=user)
        user_leads.append({
            'user': user,
            'num_leads': leads.count(),
            'leads': leads
        })
    # Use the correct related field name
    top_users = User.objects.annotate(num_leads=Count('assigned_leads')).order_by('-num_leads')[:5]

    context = {
        'total_leads': total_leads,
        'total_customers': total_customers,
        'total_users': total_users,
        'recent_leads': recent_leads,
        'overdue_leads': overdue_leads,
        'top_users': top_users,
        'user_leads': user_leads,  # Added user-wise leads
    }

    return render(request, 'admin_dashboard.html', context)


def manage_users(request):
    users = User.objects.all()

    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        action = request.POST.get('action')

        # Handle delete action
        if action == 'delete':
            user_to_delete = get_object_or_404(User, id=user_id)
            if user_to_delete.is_staff:
                messages.error(request, "Cannot delete staff users!")
            else:
                user_to_delete.delete()
                messages.success(request, "User deleted successfully.")
            return redirect('manage_users')

        # Handle deactivate action
        elif action == 'deactivate':
            user_to_deactivate = get_object_or_404(User, id=user_id)
            user_to_deactivate.is_active = False
            user_to_deactivate.save()
            messages.success(request, "User deactivated successfully.")
            return redirect('manage_users')

        # Handle activate action
        elif action == 'activate':
            user_to_activate = get_object_or_404(User, id=user_id)
            user_to_activate.is_active = True
            user_to_activate.save()
            messages.success(request, "User activated successfully.")
            return redirect('manage_users')

        # Handle promote to staff action
        elif action == 'make_staff':
            user_to_promote = get_object_or_404(User, id=user_id)
            user_to_promote.is_staff = True
            user_to_promote.save()
            messages.success(request, "User granted staff status successfully.")
            return redirect('manage_users')

        # Handle demote from staff action
        elif action == 'remove_staff':
            user_to_demote = get_object_or_404(User, id=user_id)
            user_to_demote.is_staff = False
            user_to_demote.save()
            messages.success(request, "Staff status removed successfully.")
            return redirect('manage_users')

    context = {
        'users': users
    }
    return render(request, 'manage_users.html', context)


def mark_attended(request, lead_id):
    if request.method == 'POST':
        lead = get_object_or_404(Record, id=lead_id, assigned_to=request.user)

        # Update the lead to mark the follow-up as attended
        lead.follow_up_attended = True
        lead.save()

        # Display a success message
        messages.success(request, f"Lead for {lead.client_name} marked as attended.")

        # Redirect back to the dashboard
        return redirect('home')
    else:
        # In case of a non-POST request, handle the error
        return JsonResponse({'error': 'Invalid request method.'}, status=400)
